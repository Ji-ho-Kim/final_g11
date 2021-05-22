from urllib.error import HTTPError
from selenium import webdriver
from openpyxl import Workbook, load_workbook
import datetime
import time
from selenium.webdriver.common.alert import Alert

now = datetime.datetime.now()
date = now.strftime('%Y.%m.%d')

seoul_file_path = 'C:/Users/atom2/'
seoul_file_name = 'seoul.txt'

chrome_driver_path = 'C:/Users/atom2/chromedriver.exe'

excel_file_path = 'C:/Users/atom2/'
excel_file_name = excel_file_path + date + '.xlsx'
excel_sheet_title = 'confirm'
excel_row = 2
titles = []
stars = []
types = []


def make_excel():
    work_book = Workbook()
    sheet1 = work_book.active
    sheet1.title = excel_sheet_title

    sheet1.cell(row=1, column=1).value = '위치 구'
    sheet1.cell(row=1, column=2).value = '식당이름'
    sheet1.cell(row=1, column=3).value = '평점'
    sheet1.cell(row=1, column=4).value = '유형'
    
    work_book.save(filename = excel_file_name)
    work_book.close()


def make_request():
    seoul_file = open(seoul_file_path+seoul_file_name, 'r', encoding='UTF-8')
    for seoul_code in seoul_file.readlines():
        
        for i in range (1,5):
            url = 'https://www.mangoplate.com/search/' + seoul_code + '?keyword=' + seoul_code + '&page=' + str(i)
            driver = webdriver.Chrome(chrome_driver_path)
            driver.get(url)
            time.sleep(10)
            download(driver)
            crawling(driver, seoul_code)

    seoul_file.close()

def download(driver):
    
    start = datetime.datetime.now()
    end = start + datetime.timedelta(seconds=5)
    while True:
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        time.sleep(1)
        if datetime.datetime.now() > end:
            break
    global titles
    titles = driver.find_elements_by_css_selector('body > main > article > div.column-wrapper > div > div > section > div.search-list-restaurants-inner-wrap > ul > li > div > figure > figcaption > div > a > h2')
    global stars
    stars = driver.find_elements_by_css_selector(
    'body > main > article > div.column-wrapper > div > div > section > div.search-list-restaurants-inner-wrap > ul > li > div > figure > figcaption > div > strong')
    global types
    types = driver.find_elements_by_css_selector('body > main > article > div.column-wrapper > div > div > section > div.search-list-restaurants-inner-wrap > ul > li > div > figure > figcaption > div > p.etc > span'
    )
#https://m.blog.naver.com/owl6615/221518357627
#https://galid1.tistory.com/478
#https://hello-bryan.tistory.com/194

def crawling(driver,seoul_code):
    for i in range(0,20):
        crawling_results = []
        crawling_results.append(seoul_code)
        global titles
        crawling_results.append(titles[i].text)
        print(titles[i].text)
        global stars
        crawling_results.append(stars[i].text)
        global types
        crawling_results.append(types[i].text)

        global excel_row
        insert_data_to_excel(crawling_results)
        excel_row += 1

    driver.close()
    #insert_data_to_excel(crawling_results)

def insert_data_to_excel(crawling_results):
    excel_file = load_workbook(excel_file_name)
    sheet1 = excel_file[excel_sheet_title]

    excel_column = 1;
    for data in crawling_results:
        sheet1.cell(row=excel_row, column=excel_column).value = data
        excel_column+=1

    excel_file.save(excel_file_name)
    excel_file.close()
    
make_excel()
make_request()

